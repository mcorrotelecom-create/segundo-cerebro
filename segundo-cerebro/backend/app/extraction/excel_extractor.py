"""Extracción de texto de archivos Excel, hoja por hoja.

Fase 0 extrae el contenido crudo de cada hoja en bloques de filas, con
referencia exacta a hoja + rango de filas — suficiente para que el
contenido sea buscable y citable.

Deliberadamente NO intenta en esta fase reinterpretar la jerarquía de un
presupuesto (capítulo/partida) ni reasignar el nivel/zona que viene como
texto libre dentro de una celda de comentario (ver diagnóstico de la
propuesta técnica: `NIVEL -200` aparece como fila de texto, no como
columna). Esa interpretación estructurada con verificación humana es del
patrón descrito en §03/§06 y llega en Fase 1, sobre este mismo texto ya
extraído e indexado.
"""
from app.extraction import ExtractedUnit

ROWS_PER_BLOCK = 20


def extract_excel(file_path: str) -> list[ExtractedUnit]:
    import openpyxl

    units: list[ExtractedUnit] = []
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_text: list[str] = []
        row_numbers: list[int] = []
        # Se numera con el índice propio del bucle, no con cell.row: en modo
        # read_only, openpyxl puede devolver celdas `EmptyCell` sin `.row`
        # cuando el rango declarado de la hoja no calza con su contenido
        # real (se vio en un archivo real del piloto) — confiar en el
        # índice del bucle es más robusto que confiar en el atributo.
        for row_number, row in enumerate(ws.iter_rows(), start=1):
            values = [_fmt(c.value) for c in row]
            if not any(v for v in values):
                continue
            rows_text.append(" | ".join(values))
            row_numbers.append(row_number)

            if len(rows_text) >= ROWS_PER_BLOCK:
                units.append(_flush(sheet_name, rows_text, row_numbers))
                rows_text, row_numbers = [], []

        if rows_text:
            units.append(_flush(sheet_name, rows_text, row_numbers))

    if not units:
        units.append(ExtractedUnit(text="", source_ref={"note": "hoja_vacia_o_no_legible"}))
    return units


def _flush(sheet_name: str, rows_text: list[str], row_numbers: list[int]) -> ExtractedUnit:
    text = "\n".join(rows_text)
    row_range = f"{row_numbers[0]}-{row_numbers[-1]}" if row_numbers else "?"
    return ExtractedUnit(text=text, source_ref={"sheet": sheet_name, "rows": row_range})


def _fmt(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
